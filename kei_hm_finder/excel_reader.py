'''
    File name: excel_reader.py
    Dependencies: calc_wirt.py, kei_hm.py, gauss_processor, kei_database.xlsx
    Author: Tongchen He, Nathaniel Morrison
    Date created: 1/29/2022
    Date last modified: 6/6/2022
    Python Version: 3.7.2
'''
import xlsxwriter, sys, gauss_processor, kei_hm, calc_wirt
import openpyxl

#In Python 2, which a lot of people still use because it's integrated into Linux and Sage and a host of other OSes and programs, the functionality provided by input() in
#Python 3 is instead provided by raw_input(). To make this program compatible with either version, define raw_input() to be input() if we're in Python 3, and use raw_input()
#in place of input() throughout the code.
if sys.version_info[0]==3:
    def raw_input(prompt):
        return input(prompt)

#Function to pull name and gauss code from Excel sheet
def knot_processor(cur_row, knot_workbook, knot_table, name_col, gauss_col, wirt_col):
    #Call appropriate cell and convert to string object from cell object
    name = str(knot_table.cell(cur_row,name_col).value)
    wirt_num = None
    if wirt_col > 0:
        wirt_num = int(knot_table.cell(cur_row,wirt_col).value)
    #Get rid of prefix
    name = name.lstrip("text:\'")
    #Get rid of suffix
    name=name.rstrip("\'")
    #Repeat for gauss code
    raw_gauss_code= str(knot_table.cell(cur_row,gauss_col).value)
    raw_gauss_code = raw_gauss_code.lstrip("text:\'")
    raw_gauss_code=raw_gauss_code.rstrip("\'")
    #return the name of the knot and the isolated gauss code
    return name, raw_gauss_code, wirt_num

#Function to process the keis in strings into a 2d-list
def kei_processor(raw_kei):
    result = []
    for i in range(len(raw_kei)):
        if raw_kei[i].isdigit():
            if raw_kei[i - 1].isdigit():
                result[-1]+= raw_kei[i]
            else:
                result.append(raw_kei[i])
    order = int(len(result) ** 0.5)
    kei = [[]for i in range(order)]
    for i in range(len(result)):
        kei[i//order].append(int(result[i]))
    return kei

def kei_dict_creator(file_name):
    kei_book = openpyxl.load_workbook(file_name)
    kei_sheet = kei_book.active
    kei_dict = {}
    for row_num in range(2, kei_sheet.max_row+1):
        name = str(kei_sheet.cell(row_num,1).value)
        rank = int(kei_sheet.cell(row_num,2).value)
        kei = kei_processor(str(kei_sheet.cell(row_num,3).value))
        kei_dict[name] = (kei, rank)
    kei_book.close()
    return kei_dict


#Function to create the output file
def excel_creator(excel_name):
    #Create an Excel workbook object to store excel version of output
    workbook=xlsxwriter.Workbook(excel_name)
    #Add worksheet attribute within the Excel workbook object to put the data in
    worksheet = workbook.add_worksheet()
    #Enter column titles
    worksheet.write(0, 0, 'Knot Name')
    worksheet.write(0, 1, 'Gauss Notation')
    worksheet.write(0, 2, 'Seed Strand Set')
    worksheet.write(0, 3, 'Wirtinger Number')
    worksheet.write(0, 4, 'Maps to Kei Index')
    worksheet.write(0, 5, 'Maps to Generating Set')
    #Return worksheet
    return worksheet, workbook

#Function to write data to the output file
def excel_writer(name, seed_strand_with_gauss, raw_gauss_code, wirt_num, kei_index, sym_gen_set, Knot_Number, worksheet,):
    #Write data to worksheet, with the row determined by the Knot_Number accumulator
    worksheet.write(Knot_Number, 0, str(name))
    worksheet.write(Knot_Number, 1, str(raw_gauss_code))
    worksheet.write(Knot_Number, 2, str(seed_strand_with_gauss))
    worksheet.write(Knot_Number, 3, wirt_num)
    worksheet.write(Knot_Number, 4, str(kei_index))
    worksheet.write(Knot_Number, 5, str(sym_gen_set))
    return


def excel_qnd_main(input_name):
    kei_dict = kei_dict_creator("kei_database.xlsx")
    name_col = int(input("Please enter the number of the column containing knot names in the input file (e.g. A=1, B=2, etc): "))
    gauss_col = int(input("Please enter the number of the column containing the gauss code in the input file: "))
    wirt_col = int(input("Please enter the number of the column containing the wirtinger number in the input file (-1 if doesn't have wirtinger number): "))
    start_row = int(input("Please enter the number of the row containing the first knot's information: "))
    excel_name= input("Enter desired name of excel output file (include directory and extension): ")
    knot_num = 1
    #Create a workbook object and worksheet attribute
    worksheet, workbook = excel_creator(excel_name)
    #Load data file.
    knot_workbook=openpyxl.load_workbook(input_name)
    #Call first (and only) worksheet in workbook
    knot_table=knot_workbook.active
    #Get total number of rows; since Python index starts at 0, and Excel's starts at 1, subtract 1 from the value .maxrow gives
    tot_rows=knot_table.max_row
    #Iterate through every kei in the kei dictionary
    for kei_index in kei_dict.keys():
        print("Testing kei: ", kei_index)
        # Set current row index
        cur_row = start_row
        kei = kei_dict[kei_index][0]
        rank = kei_dict[kei_index][1]
        gnr = kei_hm.generator_finder(rank, kei)
        while cur_row <= tot_rows:
            name, raw_gauss_code, wirt_num = knot_processor(cur_row, knot_workbook, knot_table, name_col, gauss_col, wirt_col)
            #run the search only when rank of the kei corresponds with wirt. num of the knot
            if wirt_num is None:
                gauss_code = gauss_processor.process_gauss_code(raw_gauss_code)
                knot_dict, seed_strand_set, wirt_num = calc_wirt.wirt_main(gauss_code)
            if wirt_num == rank:
                gauss_code = gauss_processor.process_gauss_code(raw_gauss_code)
                knot_dict, seed_strand_set, wirt_num = calc_wirt.wirt_main(gauss_code)
                hm, gen_set = kei_hm.homomorphism_finder(seed_strand_set, knot_dict, gnr, kei)
                if hm:
                    seed_strand_with_gauss = dict()
                    for strand in seed_strand_set:
                        seed_strand_with_gauss[strand] = knot_dict[strand][0]
                    # write to Excel file
                    excel_writer(name, seed_strand_with_gauss, raw_gauss_code, wirt_num, kei_index, gen_set, knot_num, worksheet)
                    # Iterate the accumulator
                    knot_num += 1
            cur_row+=1
    workbook.close()
